import xlrd
from openpyxl import *
import random
workBookName = str(input('Enter Spreadsheet name'))
sheetName = str(input('Enter tab name'))
sourceName = str(input('Enter source file name'))                
wb=Workbook()
Filename = str(workBookName+'.xlsx')
ws1 = wb.create_sheet(title=sheetName)
row = 1
with open(sourceName) as f:
    names = []
    for line in f.readlines():
        print(line)
        name = line.strip().split()
        #print(name)
        _ = ws1.cell(column=1, row = row, value = "{0}".format(line))
        row +=1
        names.append(name)
        #print(name)
print(random.choice(names))
print(len(names))
row = 1
#for name in names:
#    _ = ws1.cell(column=1, row = row, value = "{0}".format(name))
#    row +=1
wb.save(filename = Filename)
